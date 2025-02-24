import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    from tkinter import Tk
    import tkinter.filedialog
    import pandas as pd
    import openpyxl as op
    import openpyxl.styles as opstyle
    from openpyxl.styles.borders import Border, Side
    from openpyxl.utils import get_column_letter
except:
    install("pandas")
    install("openpyxl")

def pickFile():
    Tk().withdraw()
    return tkinter.filedialog.askopenfilename()

def createHeader(ws, col):
    header = ('Booth', 'Company Name', 'Day', 'Friday', 'Monday', 'Late', '@Booth', 'Shipped')
    headerByOne = 0
    # Write header titles
    while (headerByOne < len(header)):
        ws.cell(1, col).value = header[headerByOne]
        headerByOne += 1
        # Stylize cell
        ws.cell(1, col).font = opstyle.Font(bold=True, color="FFFFFF", size=16)
        ws.cell(1, col).fill = opstyle.PatternFill(start_color="000000", end_color="000000", fill_type='solid')
        col += 1

def chooseRoomColor(ws, row, col):
    ws.cell(row, col).font = opstyle.Font(size=16)
    ws.cell(row, col + 1).font = opstyle.Font(size=16)
    ws.cell(row, col + 2).font = opstyle.Font(size=16)
    # Check for missing booth placement
    if (str(ws.cell(row, col).value) == ""):
        return
    # Check if Room 1
    if (str(ws.cell(row, col).value).isdigit()):
        ws.cell(row, col).fill = opstyle.PatternFill(start_color="FEB0A8", end_color="FEB0A8", fill_type='solid')
        ws.cell(row, col + 1).fill = opstyle.PatternFill(start_color="FEB0A8", end_color="FEB0A8", fill_type='solid')
        ws.cell(row, col + 2).fill = opstyle.PatternFill(start_color="FEB0A8", end_color="FEB0A8", fill_type='solid')
    # Check if Lobby
    elif (str(ws.cell(row, col).value)[0] == 'L'):
        ws.cell(row, col).fill = opstyle.PatternFill(start_color="B3FBC2", end_color="B3FBC2", fill_type='solid')
        ws.cell(row, col + 1).fill = opstyle.PatternFill(start_color="B3FBC2", end_color="B3FBC2", fill_type='solid')
        ws.cell(row, col + 2).fill = opstyle.PatternFill(start_color="B3FBC2", end_color="B3FBC2", fill_type='solid')
    # Check if Room 2
    elif (str(ws.cell(row, col).value)[0].isalpha()):
        ws.cell(row, col).fill = opstyle.PatternFill(start_color="B3FBC2", end_color="B3FBC2", fill_type='solid')
        ws.cell(row, col + 1).fill = opstyle.PatternFill(start_color="B3FBC2", end_color="B3FBC2", fill_type='solid')
        ws.cell(row, col + 2).fill = opstyle.PatternFill(start_color="B3FBC2", end_color="B3FBC2", fill_type='solid')
    # Check if Room 3 or 4
    elif (str(ws.cell(row, col).value).startswith("3") or str(ws.cell(row, col).value).startswith("4")):
        ws.cell(row, col).fill = opstyle.PatternFill(start_color="ABE4FB", end_color="ABE4FB", fill_type='solid')
        ws.cell(row, col + 1).fill = opstyle.PatternFill(start_color="ABE4FB", end_color="ABE4FB", fill_type='solid')
        ws.cell(row, col + 2).fill = opstyle.PatternFill(start_color="ABE4FB", end_color="ABE4FB", fill_type='solid')
    # Check if Room 7 or 8
    elif (str(ws.cell(row, col).value).startswith("7") or str(ws.cell(row, col).value).startswith("8")):
        ws.cell(row, col).fill = opstyle.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
        ws.cell(row, col + 1).fill = opstyle.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
        ws.cell(row, col + 2).fill = opstyle.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')

def addBorders(ws, max_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Iterate over columns
    for col in range(1, max_col):
        # Iterate over cells in columns
        for cell in ws[get_column_letter(col)]:
            cell.border = thin_border

def main():
    # Read Excel file
    career_fair_plus_info = pd.read_excel(pickFile())
    career_fair_plus_info = career_fair_plus_info.rename(columns={"Unnamed: 1" : "Booth"})
    career_fair_plus_info = career_fair_plus_info.fillna("")
    career_fair_plus_info = career_fair_plus_info.sort_values('Employer Name').reset_index(drop=True)
    # Create matrix file
    matrixWorkbook = op.Workbook()
    # Rename sheet
    matrixWorkbook['Sheet'].title = "(Wall) Matrix"
    wallMatrix = matrixWorkbook["(Wall) Matrix"]
    # Create header
    createHeader(wallMatrix, 1)
    # Keep track of current company being parsed in Career Fair Plus Dataframe
    count = 0
    # Keep track of current column and row in new matrix
    currentRow = 2
    currentCol = 1
    # Write rows for companies in matrix
    while (count < career_fair_plus_info.shape[0]):
        # Write booth
        if "Day(s) Attending" not in career_fair_plus_info.columns:
            wallMatrix.cell(currentRow, currentCol).value = career_fair_plus_info['Unnamed: 0'][count]
            wallMatrix.column_dimensions[get_column_letter(currentCol)].width = 11.71
        else:
            wallMatrix.cell(currentRow, currentCol).value = career_fair_plus_info["Booth"][count]
            wallMatrix.column_dimensions[get_column_letter(currentCol)].width = 11.71
        currentCol += 1
        # Write company name
        wallMatrix.cell(currentRow, currentCol).value = career_fair_plus_info["Employer Name"][count]
        wallMatrix.column_dimensions[get_column_letter(currentCol)].width = 48.86
        currentCol += 1
        # Write day(s) attending
        if "Day(s) Attending" not in career_fair_plus_info.columns:
            wallMatrix.cell(currentRow, currentCol).value = ''
        elif ("Day 1" in career_fair_plus_info["Day(s) Attending"][count] and "Day 2" in career_fair_plus_info["Day(s) Attending"][count]):
            wallMatrix.cell(currentRow, currentCol).value = 'B'
        elif ("Day 1" in career_fair_plus_info["Day(s) Attending"][count]):
            wallMatrix.cell(currentRow, currentCol).value = '1'
        elif ("Day 2" in career_fair_plus_info["Day(s) Attending"][count]):
            wallMatrix.cell(currentRow, currentCol).value = '2'
        wallMatrix.column_dimensions[get_column_letter(currentCol)].width = 11.71
        # Return to booth column in matrix
        currentCol -= 2
        # Color row
        chooseRoomColor(wallMatrix, currentRow, currentCol)
        # Move to next row in matrix
        currentRow += 1
        count += 1
        if (currentRow > 67):
            currentCol += 8
            currentRow = 2
            # Create another header
            createHeader(wallMatrix, currentCol)
    # Add cell borders
    currentCol += 8
    addBorders(wallMatrix, currentCol)
    # Save Excel file
    matrixWorkbook.save('New Matrix.xlsx')

if __name__ == "__main__":
    main()